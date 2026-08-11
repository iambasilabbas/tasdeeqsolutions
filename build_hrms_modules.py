#!/usr/bin/env python3
"""
Transviti HRMS - Complete Module Builder
Builds all remaining HRMS modules
"""

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load existing workbook
wb = load_workbook('/workspace/Transviti_HRMS_v1.xlsx')

# Define styles
COLOR_HEADER = "4472C4"
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_ALT_ROW = "F2F2F2"
COLOR_CRITICAL = "FFC7CE"
COLOR_WARNING = "FFEB9C"
COLOR_SUCCESS = "C6EFCE"
COLOR_INFO = "D6EAF8"
COLOR_BORDER = "D4D4D4"

FONT_HEADER = Font(bold=True, color=COLOR_HEADER_TEXT, size=11)
FONT_NORMAL = Font(size=10)
FONT_BOLD = Font(bold=True, size=10)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

THIN_BORDER = Border(
    left=Side(style='thin', color=COLOR_BORDER),
    right=Side(style='thin', color=COLOR_BORDER),
    top=Side(style='thin', color=COLOR_BORDER),
    bottom=Side(style='thin', color=COLOR_BORDER)
)

def apply_header_style(cell):
    cell.font = FONT_HEADER
    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER

def apply_normal_style(cell, align='left'):
    cell.font = FONT_NORMAL
    cell.border = THIN_BORDER
    if align == 'center':
        cell.alignment = ALIGN_CENTER
    elif align == 'right':
        cell.alignment = ALIGN_RIGHT
    else:
        cell.alignment = ALIGN_LEFT

def apply_alt_row_style(cell):
    cell.font = FONT_NORMAL
    cell.fill = PatternFill(start_color=COLOR_ALT_ROW, end_color=COLOR_ALT_ROW, fill_type='solid')
    cell.border = THIN_BORDER
    cell.alignment = ALIGN_LEFT

def apply_critical_style(cell):
    cell.font = FONT_NORMAL
    cell.fill = PatternFill(start_color=COLOR_CRITICAL, end_color=COLOR_CRITICAL, fill_type='solid')
    cell.border = THIN_BORDER
    cell.alignment = ALIGN_LEFT

def apply_warning_style(cell):
    cell.font = FONT_NORMAL
    cell.fill = PatternFill(start_color=COLOR_WARNING, end_color=COLOR_WARNING, fill_type='solid')
    cell.border = THIN_BORDER
    cell.alignment = ALIGN_LEFT

def apply_success_style(cell):
    cell.font = FONT_NORMAL
    cell.fill = PatternFill(start_color=COLOR_SUCCESS, end_color=COLOR_SUCCESS, fill_type='solid')
    cell.border = THIN_BORDER
    cell.alignment = ALIGN_LEFT

# ============================================================
# SHEET 4: Employee Documents
# ============================================================
print("Creating Employee Documents sheet...")
ws_docs = wb.create_sheet('Employee Documents')

doc_headers = [
    'Employee ID', 'Employee Name', 'Department', 'CNIC Status', 'CV Status',
    'Offer Letter Status', 'Educational Docs Status', 'Experience Letter Status',
    'NDA Status', 'Passport Status', 'Bank Details Status', 'Emergency Contact Status',
    'Medical Certificate Status', 'Police Certificate Status', 'Total Documents Required',
    'Documents Submitted', 'Compliance %', 'Compliance Status', 'Last Updated', 'Notes'
]

for col, header in enumerate(doc_headers, 1):
    cell = ws_docs.cell(row=1, column=col, value=header)
    apply_header_style(cell)

# Add formulas for existing employees (rows 2-40)
for row in range(2, 41):
    # Employee ID lookup
    ws_docs.cell(row=row, column=1, value=f"=Employee_Master!A{row}")
    # Employee Name lookup
    ws_docs.cell(row=row, column=2, value=f"=Employee_Master!B{row}")
    # Department lookup
    ws_docs.cell(row=row, column=3, value=f"=Employee_Master!P{row}")
    
    # Document status columns (defaults to Pending)
    for col in range(4, 15):
        ws_docs.cell(row=row, column=col, value='Pending')
    
    # Total documents required (configurable - default 10)
    ws_docs.cell(row=row, column=15, value=10)
    
    # Documents submitted - will count non-Pending statuses
    ws_docs.cell(row=row, column=16, value=f"=COUNTIF(D{row}:N{row},\"Submitted\")")
    
    # Compliance percentage
    ws_docs.cell(row=row, column=17, value=f"=IF(O{row}>0,Q{row}/O{row}*100,0)")
    
    # Compliance status formula
    ws_docs.cell(row=row, column=18, value=f'=IF(Q{row}>=10,"Complete",IF(Q{row}>=5,"Partial","Incomplete"))')
    
    # Last updated
    ws_docs.cell(row=row, column=19, value='')
    # Notes
    ws_docs.cell(row=row, column=20, value='')
    
    # Apply styles
    for col in range(1, 21):
        cell = ws_docs.cell(row=row, column=col)
        if col <= 3 or col >= 19:
            apply_normal_style(cell)
        elif col == 18:
            apply_normal_style(cell, 'center')
        else:
            apply_normal_style(cell, 'center')

# Set column widths
for col, width in enumerate([12, 25, 20, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 12, 15, 15, 30], 1):
    ws_docs.column_dimensions[get_column_letter(col)].width = width

print("Created Employee Documents sheet")

# ============================================================
# SHEET 5: Onboarding Tracker
# ============================================================
print("Creating Onboarding Tracker sheet...")
ws_onboard = wb.create_sheet('Onboarding Tracker')

onboard_headers = [
    'Employee ID', 'Employee Name', 'DOJ', 'Department', 'Designation',
    '1. Record Created', '2. ID Assigned', '3. CNIC Submitted', '4. CV Submitted',
    '5. Educational Docs', '6. Offer Letter Signed', '7. Personal Info Complete',
    '8. Emergency Contact', '9. Bank Info Complete', '10. Manager Assigned',
    '11. Department Assigned', '12. Designation Assigned', '13. Probation Recorded',
    '14. Orientation Done', '15. Assets Issued', '16. System Access', '17. HR Review',
    'Total Tasks', 'Completed Tasks', 'Completion %', 'Onboarding Status', 'Start Date',
    'Target End Date', 'Actual End Date', 'Blocked Reason', 'Notes'
]

for col, header in enumerate(onboard_headers, 1):
    cell = ws_onboard.cell(row=1, column=col, value=header)
    apply_header_style(cell)

# Add formulas for existing employees
for row in range(2, 41):
    ws_onboard.cell(row=row, column=1, value=f"=Employee_Master!A{row}")
    ws_onboard.cell(row=row, column=2, value=f"=Employee_Master!B{row}")
    ws_onboard.cell(row=row, column=3, value=f"=Employee_Master!M{row}")
    ws_onboard.cell(row=row, column=4, value=f"=Employee_Master!P{row}")
    ws_onboard.cell(row=row, column=5, value=f"=Employee_Master!Q{row}")
    
    # Task checkboxes (Y/N) - default empty
    for col in range(6, 23):
        ws_onboard.cell(row=row, column=col, value='')
    
    # Total tasks
    ws_onboard.cell(row=row, column=23, value=17)
    
    # Completed tasks count
    ws_onboard.cell(row=row, column=24, value='=COUNTIF(F2:V2,"Y")')
    
    # Completion percentage
    ws_onboard.cell(row=row, column=25, value=f'=IF(W{row}>0,X{row}/W{row}*100,0)')
    
    # Onboarding status
    ws_onboard.cell(row=row, column=26, value=f'=IF(Y{row}=100,"Completed",IF(Y{row}>0,"In Progress","Not Started"))')
    
    # Dates and notes
    for col in range(27, 31):
        ws_onboard.cell(row=row, column=col, value='')
    
    for col in range(1, 31):
        apply_normal_style(ws_onboard.cell(row=row, column=col))

# Column widths
ws_onboard.column_dimensions['A'].width = 12
ws_onboard.column_dimensions['B'].width = 25
ws_onboard.column_dimensions['C'].width = 12
ws_onboard.column_dimensions['D'].width = 20
ws_onboard.column_dimensions['E'].width = 25
for i in range(6, 23):
    ws_onboard.column_dimensions[get_column_letter(i)].width = 4
for i, w in enumerate([10, 10, 12, 18, 12, 15, 15, 20, 30], 23):
    ws_onboard.column_dimensions[get_column_letter(i)].width = w

print("Created Onboarding Tracker sheet")

# ============================================================
# SHEET 6: Offboarding Tracker
# ============================================================
print("Creating Offboarding Tracker sheet...")
ws_offboard = wb.create_sheet('Offboarding Tracker')

offboard_headers = [
    'Employee ID', 'Employee Name', 'Department', 'Designation',
    'Resignation/Termination Date', 'Reason', 'Notice Period Days', 'Last Working Day',
    '1. Handover Complete', '2. Assets Returned', '3. IT Clearance', '4. Email Clearance',
    '5. Finance Clearance', '6. Manager Clearance', '7. HR Clearance', '8. Exit Interview',
    '9. Experience Letter', '10. Relieving Letter', '11. Final Settlement',
    'Total Tasks', 'Completed Tasks', 'Completion %', 'Offboarding Status',
    'Final Approval', 'Approved By', 'Approval Date', 'Blocked Reason', 'Notes'
]

for col, header in enumerate(offboard_headers, 1):
    cell = ws_offboard.cell(row=1, column=col, value=header)
    apply_header_style(cell)

# Add rows for exited employees (currently none, but structure ready)
for row in range(2, 41):
    ws_offboard.cell(row=row, column=1, value=f"=Employee_Master!A{row}")
    ws_offboard.cell(row=row, column=2, value=f"=Employee_Master!B{row}")
    ws_offboard.cell(row=row, column=3, value=f"=Employee_Master!P{row}")
    ws_offboard.cell(row=row, column=4, value=f"=Employee_Master!Q{row}")
    
    for col in range(5, 29):
        ws_offboard.cell(row=row, column=col, value='')
    
    for col in range(1, 29):
        apply_normal_style(ws_offboard.cell(row=row, column=col))

# Column widths
for i, w in enumerate([12, 25, 20, 25, 15, 30, 12, 15, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 10, 10, 12, 18, 15, 20, 15, 20, 30], 1):
    ws_offboard.column_dimensions[get_column_letter(i)].width = w

print("Created Offboarding Tracker sheet")

# Save intermediate
wb.save('/workspace/Transviti_HRMS_v2.xlsx')
print("Saved intermediate workbook (v2)")
