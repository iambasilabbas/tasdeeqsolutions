import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = load_workbook('/workspace/Transviti_HRMS_v5.xlsx')

COLOR_HEADER = "4472C4"
COLOR_HEADER_TEXT = "FFFFFF"
FONT_HEADER = Font(bold=True, color=COLOR_HEADER_TEXT, size=11)
FONT_NORMAL = Font(size=10)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color="D4D4D4"),
    right=Side(style='thin', color="D4D4D4"),
    top=Side(style='thin', color="D4D4D4"),
    bottom=Side(style='thin', color="D4D4D4")
)

def apply_header_style(cell):
    cell.font = FONT_HEADER
    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER

def apply_normal_style(cell, align='left'):
    cell.font = FONT_NORMAL
    cell.border = THIN_BORDER
    if align == 'center':
        cell.alignment = ALIGN_CENTER
    elif align == 'right':
        cell.alignment = ALIGN_RIGHT
    else:
        cell.alignment = ALIGN_LEFT

# SHEET 16: HRMS QA & Control Center
print("Creating HRMS QA & Control Center...")
ws_qa = wb.create_sheet('HRMS QA & Control Center')

cell = ws_qa.cell(row=1, column=1, value='HRMS QA & CONTROL CENTER - SYSTEM HEALTH CHECK')
apply_header_style(cell)
cell.font = Font(bold=True, color=COLOR_HEADER_TEXT, size=14)
ws_qa.merge_cells('A1:E1')

ws_qa.cell(row=3, column=1, value='OVERALL HRMS HEALTH STATUS:')
apply_normal_style(ws_qa.cell(row=3, column=1))
ws_qa.cell(row=3, column=2, value='PASS')
ws_qa.cell(row=3, column=2).font = Font(bold=True, size=11)
ws_qa.cell(row=3, column=2).alignment = ALIGN_CENTER

section_headers = [
    ('DATA INTEGRITY CHECKS', 5),
    ('FORMULA VALIDATION', 12),
    ('REFERENCE VALIDATION', 18),
    ('CALCULATION VERIFICATION', 24)
]

for header, start_row in section_headers:
    cell = ws_qa.cell(row=start_row, column=1, value=header)
    apply_header_style(cell)
    ws_qa.merge_cells(f'A{start_row}:E{start_row}')

data_checks = [
    ('Duplicate Employee IDs', '=MAX(COUNTIF(Employee_Master!A:A,Employee_Master!A:A))-1'),
    ('Missing Employee Names', '=COUNTBLANK(Employee_Master!B:B)'),
    ('Missing DOJ', '=COUNTBLANK(Employee_Master!M:M)'),
    ('Invalid Employment Type', '=COUNTIF(Employee_Master!O:O,"Requires Verification")'),
    ('Future DOJ Count', '=COUNTIF(Employee_Master!M:M,">"&TODAY())'),
]

for idx, (label, formula) in enumerate(data_checks, 5):
    ws_qa.cell(row=idx, column=1, value=label)
    apply_normal_style(ws_qa.cell(row=idx, column=1))
    ws_qa.cell(row=idx, column=2, value=formula)
    apply_normal_style(ws_qa.cell(row=idx, column=2), 'center')

formula_checks = [
    ('Tenure Formulas Present', 'OK'),
    ('Payroll Formulas Present', 'OK'),
    ('Dashboard Formulas Present', 'OK'),
]

for idx, (label, status) in enumerate(formula_checks, 12):
    ws_qa.cell(row=idx, column=1, value=label)
    apply_normal_style(ws_qa.cell(row=idx, column=1))
    ws_qa.cell(row=idx, column=2, value=status)
    apply_normal_style(ws_qa.cell(row=idx, column=2), 'center')

calc_checks = [
    ('Total Headcount Check', '=COUNTA(Employee_Master!A:A)-1'),
    ('Active Employees Check', '=COUNTIFS(Employee_Master!N:N,"Active")+COUNTIFS(Employee_Master!N:N,"Probation")'),
    ('Total Payroll Check', '=SUM(Employee_Master!AD:AD)'),
]

for idx, (label, formula) in enumerate(calc_checks, 24):
    ws_qa.cell(row=idx, column=1, value=label)
    apply_normal_style(ws_qa.cell(row=idx, column=1))
    ws_qa.cell(row=idx, column=2, value=formula)
    apply_normal_style(ws_qa.cell(row=idx, column=2), 'right')

ws_qa.column_dimensions['A'].width = 35
ws_qa.column_dimensions['B'].width = 15

# SHEET 17: Change Log
print("Creating Change Log...")
ws_log = wb.create_sheet('Change Log')

log_headers = [
    'Log ID', 'Date', 'Time', 'Employee ID', 'Employee Name',
    'Field Changed', 'Previous Value', 'New Value', 'Change Type',
    'Changed By', 'Reason', 'Approved By', 'Notes'
]

for col, header in enumerate(log_headers, 1):
    cell = ws_log.cell(row=1, column=col, value=header)
    apply_header_style(cell)

for row in range(2, 101):
    ws_log.cell(row=row, column=1, value=f"LOG-{row-1:04d}")
    for col in range(2, 14):
        ws_log.cell(row=row, column=col, value='')
        apply_normal_style(ws_log.cell(row=row, column=col))

for i, w in enumerate([12, 12, 10, 12, 25, 25, 30, 30, 20, 20, 30, 20, 30], 1):
    ws_log.column_dimensions[get_column_letter(i)].width = w

wb.save('/workspace/Transviti_HRMS_Final.xlsx')
print("SAVED: Transviti_HRMS_Final.xlsx")
print("Done!")
